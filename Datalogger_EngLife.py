"""
Created on Tue Sep 17 23:49:15 2024
pyinstaller --onefile --windowed --icon=englife.ico Datalogger_EngLife.py

@author: roger
"""


import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
from collections import defaultdict
import os
import mplcursors
from tkcalendar import DateEntry  # Importa o DateEntry do tkcalendar
from fpdf import FPDF
from openpyxl.styles import Font, PatternFill
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import babel
from babel.dates import format_date
from datetime import date
import tkinter as tk
from tkcalendar import DateEntry
from babel.dates import format_date
from babel.numbers import format_number
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import statsmodels.api as sm
import requests

# Definindo o locale para pt_BR
locale = babel.Locale('pt', 'BR')

def ler_dados(arquivo):
    """Lê os dados do arquivo e retorna uma lista de dicionários, ajustando os nomes das três últimas colunas do cabeçalho."""
    dados = []
    
    with open(arquivo, 'r', encoding='utf-8') as file:
        # Lê o cabeçalho e ajusta os nomes das três últimas colunas
        cabecalho = next(file).strip().split('\t')
        if len(cabecalho) >= 3:
            cabecalho[-3:] = ['Temperatura da Água', 'Temperatura da Estufa', 'Temperatura Externa']
        
        # Lê os dados e os armazena em uma lista de dicionários
        for linha in file:
            colunas = linha.strip().split('\t')
            dados.append(dict(zip(cabecalho, colunas)))
    
    return dados, cabecalho
def periodo_coleta(dados):
    """Determina o período de coleta dos dados."""
    # Ajustando para o formato 'dd/mm/aaaa'
    data_inicial = datetime.strptime(dados[0]['Data'], '%d/%m/%Y')
    data_final = datetime.strptime(dados[-1]['Data'], '%d/%m/%Y')
    
    # Formatando para mostrar apenas a data sem hora
    data_inicial = data_inicial.strftime('%d/%m/%Y')
    data_final = data_final.strftime('%d/%m/%Y')
    
    print("Data inicial:", data_inicial)
    print("Data final:", data_final)
    
    return data_inicial, data_final
def filtrar_dados_por_periodo(dados, data_inicio, data_fim):
    """Filtra os dados para o período especificado (inclusivo)."""
    dados_filtrados = []
    
    data_inicio = datetime.strptime(data_inicio, '%d-%m-%Y').date()
    data_fim = datetime.strptime(data_fim, '%d-%m-%Y').date()
    
    for dado in dados:
        data_dado = datetime.strptime(dado['Data'], '%d/%m/%Y').date()
        if data_inicio <= data_dado <= data_fim:
            dados_filtrados.append(dado)
    
    return dados_filtrados
def calcular_intervalo_entre_amostras(dados):
    """
    Calcula o intervalo médio de tempo em minutos entre amostras consecutivas nos dados.

    Parâmetros:
    - dados: lista de dicionários contendo 'Data' e 'Hora' no formato '%d/%m/%Y' e '%H:%M:%S'.

    Retorna:
    - Intervalo médio de tempo entre as amostras em minutos.
    """
    intervalos = []
    
    for i in range(1, len(dados)):
        data_hora_atual = datetime.strptime(f"{dados[i]['Data']} {dados[i]['Hora']}", '%d/%m/%Y %H:%M:%S')
        data_hora_anterior = datetime.strptime(f"{dados[i-1]['Data']} {dados[i-1]['Hora']}", '%d/%m/%Y %H:%M:%S')
        
        # Calcula o intervalo de tempo em minutos
        intervalo = (data_hora_atual - data_hora_anterior).total_seconds() / 60
        intervalos.append(intervalo)
    
    if intervalos:
        # Calcula o intervalo médio
        intervalo_medio = sum(intervalos) / len(intervalos)
        return intervalo_medio
    else:
        return None  # Retorna None se não houver dados suficientes
def calcular_amplitude_termica(dados, sensores):
    """
    Calcula a amplitude térmica de cada dia e retorna:
    - A data com a maior amplitude térmica do sensor de temperatura da água.
    - Temperaturas máximas, mínimas e variação para cada sensor nesse dia.
    """
    # Organiza as temperaturas por data e sensor
    temperaturas = defaultdict(lambda: {sensor: [] for sensor in sensores})
    
    for dado in dados:
        data = dado['Data']
        for sensor in sensores:
            temperaturas[data][sensor].append(float(dado[sensor]))
    
    # Calcula a amplitude térmica diária para cada sensor
    amplitudes = {}
    for data, sensores_temperaturas in temperaturas.items():
        amplitude_total = 0
        for sensor, temps in sensores_temperaturas.items():
            if temps:
                amplitude = max(temps) - min(temps)
                amplitude_total = max(amplitude_total, amplitude)
        amplitudes[data] = amplitude_total
    
    # Identifica o dia com a maior amplitude térmica para o sensor de temperatura da água
    sensor_agua = sensores[0]
    temperaturas_agua = temperaturas
    amplitudes_agua = {}
    for data, sensores_temperaturas in temperaturas_agua.items():
        if sensor_agua in sensores_temperaturas and temperaturas_agua[data][sensor_agua]:
            amplitude_agua = max(temperaturas_agua[data][sensor_agua]) - min(temperaturas_agua[data][sensor_agua])
            amplitudes_agua[data] = amplitude_agua
    
    dia_max_amplitude_agua = max(amplitudes_agua, key=amplitudes_agua.get)
    
    # Obtém as temperaturas e a variação para o dia com a maior amplitude do sensor de água
    dados_max_amplitude = temperaturas[dia_max_amplitude_agua]
    resultados = {}
    
    for sensor, temps in dados_max_amplitude.items():
        if temps:
            temp_max = round(max(temps), 2)
            temp_min = round(min(temps), 2)
            var = round(temp_max - temp_min, 2)
            resultados[sensor] = {
                'temp_max': temp_max,
                'temp_min': temp_min,
                'variacao': var
            }
    
    return dia_max_amplitude_agua, resultados
def calcular_variacao(dados, sensores, dia_max_amplitude):
    """Calcula a variação entre amostras para um dia específico e retorna o pico de variação com aumento ou redução, limitado a duas casas decimais."""
    variacoes = {sensor: [] for sensor in sensores}
    prev_dado = {sensor: None for sensor in sensores}
    
    # Filtra os dados para o dia específico
    dados_filtrados = [dado for dado in dados if dado['Data'] == dia_max_amplitude]

    for dado in dados_filtrados:
        hora = dado['Hora']
        for sensor in sensores:
            temp = float(dado[sensor])
            if prev_dado[sensor] is not None:
                variacao = temp - prev_dado[sensor]
                variacao = round(variacao, 2)  # Limita a variação a duas casas decimais
                variacoes[sensor].append((variacao, hora))
            prev_dado[sensor] = temp
    
    picos_variacao = {}
    
    for sensor, var_list in variacoes.items():
        if var_list:
            pico = max(var_list, key=lambda x: abs(x[0]))  # Encontra a variação com maior valor absoluto
            tipo_variacao = 'Aumento' if pico[0] > 0 else 'Queda'
            picos_variacao[sensor] = (round(pico[0], 2), tipo_variacao, pico[1])  # Limita o pico a duas casas decimais
    
    return picos_variacao
def calcular_estatisticas_temperaturas(dados_filtrados, sensores):
    """
    Calcula a média das temperaturas do período, a média das temperaturas mais altas e baixas de cada dia,
    e a média da amplitude térmica para cada sensor.
    
    Parâmetros:
        - dados_filtrados: lista de dicionários contendo os dados filtrados por data
        - sensores: lista com os nomes dos sensores
    
    Retorna:
        - Um dicionário contendo as médias calculadas para cada sensor
    """
    estatisticas_sensores = {}

    # Organizar os dados por dia
    dados_por_dia = {}
    for dado in dados_filtrados:
        data = dado['Data']
        if data not in dados_por_dia:
            dados_por_dia[data] = []
        dados_por_dia[data].append(dado)

    # Calcular as médias por sensor
    for sensor in sensores:
        temperaturas_periodo = []
        temperaturas_max_dias = []
        temperaturas_min_dias = []
        amplitudes_dias = []

        for data, dados_dia in dados_por_dia.items():
            # Extrair as temperaturas do sensor para o dia
            temperaturas_dia = [float(dado[sensor]) for dado in dados_dia]

            # Calcular temperatura máxima, mínima e amplitude do dia
            max_temp_dia = max(temperaturas_dia)
            min_temp_dia = min(temperaturas_dia)
            amplitude_dia = max_temp_dia - min_temp_dia

            # Armazenar para cálculos de médias
            temperaturas_max_dias.append(max_temp_dia)
            temperaturas_min_dias.append(min_temp_dia)
            amplitudes_dias.append(amplitude_dia)
            temperaturas_periodo.extend(temperaturas_dia)

        # Calcular as médias para o sensor
        media_temperatura_periodo = sum(temperaturas_periodo) / len(temperaturas_periodo) if temperaturas_periodo else 0
        media_temperatura_max = sum(temperaturas_max_dias) / len(temperaturas_max_dias) if temperaturas_max_dias else 0
        media_temperatura_min = sum(temperaturas_min_dias) / len(temperaturas_min_dias) if temperaturas_min_dias else 0
        media_amplitude = sum(amplitudes_dias) / len(amplitudes_dias) if amplitudes_dias else 0

        # Armazenar os resultados no dicionário
        estatisticas_sensores[sensor] = {
            'media_temperatura_periodo': media_temperatura_periodo,
            'media_temperatura_max': media_temperatura_max,
            'media_temperatura_min': media_temperatura_min,
            'media_amplitude': media_amplitude
        }

    return estatisticas_sensores
def calcular_dia_mais_quente(dados, sensores):
    """
    Calcula o dia mais quente para cada sensor e retorna:
    - A data com a maior temperatura média para cada sensor.
    - Temperaturas máximas, mínimas e variação para cada sensor nesse dia.
    """
    # Organiza as temperaturas por data e sensor
    temperaturas = defaultdict(lambda: {sensor: [] for sensor in sensores})
    
    for dado in dados:
        data = dado['Data']
        for sensor in sensores:
            temperaturas[data][sensor].append(float(dado[sensor]))
    
    # Calcula a temperatura média diária para cada sensor
    medias_temperaturas = {}
    for data, sensores_temperaturas in temperaturas.items():
        medias_temperaturas[data] = {}
        for sensor, temps in sensores_temperaturas.items():
            if temps:
                medias_temperaturas[data][sensor] = sum(temps) / len(temps)
    
    # Identifica o dia com a maior temperatura média para cada sensor
    dia_max_temp = max(medias_temperaturas, key=lambda data: max(medias_temperaturas[data].values()))
    
    # Obtém as temperaturas e a variação para o dia mais quente
    resultados = {}
    for sensor, temps in temperaturas[dia_max_temp].items():
        if temps:
            temp_max = round(max(temps), 2)
            temp_min = round(min(temps), 2)
            var = round(temp_max - temp_min, 2)
            resultados[sensor] = {
                'temp_max': temp_max,
                'temp_min': temp_min,
                'variacao': var
            }
    
    return dia_max_temp, resultados
def calcular_dia_mais_frio(dados, sensores):
    """
    Calcula o dia mais frio para cada sensor e retorna:
    - A data com a menor temperatura média para cada sensor.
    - Temperaturas máximas, mínimas e variação para cada sensor nesse dia.
    """
    # Organiza as temperaturas por data e sensor
    temperaturas = defaultdict(lambda: {sensor: [] for sensor in sensores})
    
    for dado in dados:
        data = dado['Data']
        for sensor in sensores:
            temperaturas[data][sensor].append(float(dado[sensor]))
    
    # Calcula a temperatura média diária para cada sensor
    medias_temperaturas = {}
    for data, sensores_temperaturas in temperaturas.items():
        medias_temperaturas[data] = {}
        for sensor, temps in sensores_temperaturas.items():
            if temps:
                medias_temperaturas[data][sensor] = sum(temps) / len(temps)
    
    # Identifica o dia com a menor temperatura média para cada sensor
    dia_min_temp = min(medias_temperaturas, key=lambda data: min(medias_temperaturas[data].values()))
    
    # Obtém as temperaturas e a variação para o dia mais frio
    resultados = {}
    for sensor, temps in temperaturas[dia_min_temp].items():
        if temps:
            temp_max = round(max(temps), 2)
            temp_min = round(min(temps), 2)
            var = round(temp_max - temp_min, 2)
            resultados[sensor] = {
                'temp_max': temp_max,
                'temp_min': temp_min,
                'variacao': var
            }
    
    return dia_min_temp, resultados
def definir_limites(temperatura_ar_min, temperatura_ar_max, temperatura_agua_min, temperatura_agua_max, variacao_agua_max):
    """
    Define os limites de temperatura para a análise.

    Parâmetros:
    - temperatura_ar_min: limite mínimo da temperatura do ar.
    - temperatura_ar_max: limite máximo da temperatura do ar.
    - temperatura_agua_min: limite mínimo da temperatura da água.
    - temperatura_agua_max: limite máximo da temperatura da água.
    - variacao_agua_max: limite máximo de variação da temperatura da água.

    Retorna:
    - Um dicionário contendo os limites definidos.
    """
    return {
        'Temperatura Ar': {
            'minima': temperatura_ar_min,
            'maxima': temperatura_ar_max
        },
        'Temperatura Água': {
            'minima': temperatura_agua_min,
            'maxima': temperatura_agua_max
        },
        'Variação Água': {
            'maxima': variacao_agua_max
        }
    }
def avaliar_condicoes(dados, sensores, limites):
    """
    Avalia as condições de temperatura para cada dia e retorna os dias em que as condições não foram atendidas
    junto com os parâmetros específicos que estão fora dos limites, incluindo uma breve explicação sobre o problema.
    
    Parâmetros:
    - dados: lista de dicionários contendo 'Data' e valores dos sensores.
    - sensores: lista com os nomes dos sensores em ordem [sensor_agua, sensor_estufa, sensor_externa].

    Retorna:
    - Lista de tuplas contendo a data e os motivos (parâmetros fora dos limites).
    """
    
    # Organizando as temperaturas por data e sensor
    temperaturas = defaultdict(lambda: {sensor: [] for sensor in sensores})
    
    for dado in dados:
        data = dado['Data']
        for sensor in sensores:
            temperaturas[data][sensor].append(float(dado[sensor]))

    # Armazenando os dias em que as condições não foram atendidas
    dias_problemas = []

    # Avaliando as condições para cada dia
    for data, sensores_temperaturas in temperaturas.items():
        problema = []
        
        # Sensor de água (sempre o primeiro sensor da lista)
        sensor_agua = sensores[0]
        temp_agua = temperaturas[data][sensor_agua]
        media_temp_agua = sum(temp_agua) / len(temp_agua)
        variacao_agua = max(temp_agua) - min(temp_agua)

        if media_temp_agua < limites['Temperatura Água']['minima']:
            problema.append(f"Temperatura da Água abaixo do mínimo permitido (Média do dia: {media_temp_agua:.2f}°C, Mínima recomendado: {limites['Temperatura Água']['minima']}°C).\n "
                            "A temperatura da água está muito baixa, o que pode impactar a atividade e o metabolismo dos peixes.")
        if media_temp_agua > limites['Temperatura Água']['maxima']:
            problema.append(f"Temperatura da Água acima do máximo permitido (Média do dia: {media_temp_agua:.2f}°C, Máxima recomendado: {limites['Temperatura Água']['maxima']}°C).\n "
                            "A temperatura elevada da água pode causar estresse térmico nos peixes.")
        if variacao_agua > limites['Variação Água']['maxima']:
            problema.append(f"Variação da Temperatura da Água acima do limite permitido (Variação do dia: {variacao_agua:.2f}°C, Variação Máxima recomendada: {limites['Variação Água']['maxima']}°C).\n "
                            "A temperatura está esta instável, isso estressar os animais.\n Fique atendo ao menejo da cortina e evite deixar a porta aberta.")

        # Sensor de ar (considerando o sensor da estufa, que é sempre o segundo da lista)
        sensor_estufa = sensores[1]
        temp_estufa = temperaturas[data][sensor_estufa]
        media_temp_estufa = sum(temp_estufa) / len(temp_estufa)

        if media_temp_estufa < limites['Temperatura Ar']['minima']:
            problema.append(f"{sensor_estufa}: Temperatura da Estufa abaixo do mínimo permitido (Média do dia: {media_temp_estufa:.2f}°C, Mínima recomendado: {limites['Temperatura Ar']['minima']}°C. \n) "
                            "A temperatura da estufa está muito baixa, o que pode causar desconforto térmico e estresse.")
        elif media_temp_estufa > limites['Temperatura Ar']['maxima']:
            problema.append(f"{sensor_estufa}: Temperatura da Estufa acima do máximo permitido (Média do dia: {media_temp_estufa:.2f}°C, Máxima recomendado: {limites['Temperatura Ar']['maxima']}°C. \n) "
                            "A temperatura da estufa está muito alta, podendo resultar em superaquecimento.")

        # Se houver problemas, armazenamos a data e os motivos
        if problema:
            dias_problemas.append((data, problema))

    return dias_problemas
def avaliar_impacto_estufa(dados, sensores, max_lag=24):
    """
    Avalia quanto a temperatura externa (Sensor 3) precisa variar para alterar a temperatura da água (Sensor 1)
    e a temperatura interna (Sensor 2), e o tempo necessário para essa alteração se refletir na temperatura da água.

    Parâmetros:
    - dados: lista de dicionários contendo 'Data', 'Hora', e leituras dos sensores 1, 2 e 3.
    - sensores: lista com os nomes dos sensores em ordem [sensor_agua, sensor_interna, sensor_externa].
    - max_lag: número máximo de períodos de defasagem a serem testados.

    Retorna:
    - Quanto a temperatura externa precisa variar para alterar 1°C na água.
    - Quanto a temperatura externa precisa variar para alterar 1°C na temperatura interna.
    - O tempo necessário para a alteração da temperatura externa se refletir na água.
    - Intervalo médio de tempo entre amostras.
    """
    sensor_agua = sensores[0]
    sensor_interna = sensores[1]
    sensor_externa = sensores[2]
    
    # Converte os dados em um DataFrame
    df = pd.DataFrame(dados)
    
    # Certifica-se de que as colunas estejam no formato correto (float para as leituras de temperatura)
    df[sensor_agua] = df[sensor_agua].astype(float)
    df[sensor_interna] = df[sensor_interna].astype(float)
    df[sensor_externa] = df[sensor_externa].astype(float)

    # Calcula as variações entre amostras consecutivas para cada sensor
    df['Var_Agua'] = df[sensor_agua].diff()
    df['Var_Interna'] = df[sensor_interna].diff()
    df['Var_Externa'] = df[sensor_externa].diff()
    
    # Remove as linhas com valores NaN resultantes do cálculo das diferenças
    df = df.dropna()

    # Verifica se o DataFrame resultante está vazio
    if df.empty:
        raise ValueError("Os dados processados resultaram em um DataFrame vazio. Verifique se há valores suficientes após o cálculo das diferenças.")

    melhor_r2 = -float('inf')
    melhor_lag = 0
    modelo_agua_final = None

    # Testa defasagens de 0 a max_lag
    for lag in range(max_lag + 1):
        df[f'Var_Externa_Lag{lag}'] = df['Var_Externa'].shift(lag)
        
        # Regressão: Variação da temperatura da água em relação à variação externa defasada e interna
        X_agua = df[['Var_Externa_Lag' + str(lag), 'Var_Interna']].dropna()
        y_agua = df['Var_Agua'][X_agua.index]  # Alinha a variável dependente

        # Verifica se X_agua e y_agua têm valores suficientes para a regressão
        if len(X_agua) > 1:  # Necessário pelo menos 2 pontos para ajuste
            X_agua = sm.add_constant(X_agua)  # Adiciona a constante (intercepto) ao modelo
            
            # Adiciona tratamento de exceção para a regressão
            try:
                modelo_agua = sm.OLS(y_agua, X_agua).fit()

                # Avalia o R² do modelo
                if modelo_agua.rsquared > melhor_r2:
                    melhor_r2 = modelo_agua.rsquared
                    melhor_lag = lag
                    modelo_agua_final = modelo_agua

            except Exception as e:
                print(f"Erro ao ajustar o modelo para lag {lag}: {e}")
        else:
            print(f"Defasagem {lag} ignorada por falta de dados suficientes.")

    # Regressão: Variação da temperatura interna em relação à variação externa (sem lag)
    X_interna = df[['Var_Externa']].dropna()
    y_interna = df['Var_Interna'][X_interna.index]  # Alinha a variável dependente

    # Verifica se X_interna e y_interna têm valores suficientes
    if len(X_interna) > 1:
        X_interna = sm.add_constant(X_interna)  # Adiciona a constante (intercepto) ao modelo
        
        # Adiciona tratamento de exceção para a regressão
        try:
            modelo_interna = sm.OLS(y_interna, X_interna).fit()
        except Exception as e:
            raise ValueError(f"Erro ao ajustar o modelo para a temperatura interna: {e}")
    else:
        raise ValueError("Não há dados suficientes para a regressão da temperatura interna.")

    # Calcula o intervalo médio entre as amostras
    intervalo_medio = calcular_intervalo_entre_amostras(dados)

    # Coeficientes para determinar quanto a temperatura externa precisa variar para afetar a água e a interna
    if modelo_agua_final is None:
        raise ValueError("Não foi possível ajustar um modelo adequado para a variação da temperatura da água.")
    
    coef_agua = abs(modelo_agua_final.params[f'Var_Externa_Lag{melhor_lag}'])  # Usar o valor absoluto
    coef_interna = abs(modelo_interna.params['Var_Externa'])  # Usar o valor absoluto

    # Calcula quanto a temperatura externa precisa variar para mudar 1°C na água e na interna
    variacao_necessaria_agua = abs(1 / coef_agua)  # Garante que o valor seja positivo
    variacao_necessaria_interna = abs(1 / coef_interna)  # Garante que o valor seja positivo

    # Calcula o tempo necessário para a alteração da temperatura externa afetar a água
    tempo_reflexao_agua = melhor_lag * intervalo_medio
    # Converte para horas se o tempo for maior que 60 minutos
    if tempo_reflexao_agua > 60:
        tempo_reflexao_agua /= 60  # Converte para horas
        unidade_tempo = "horas"
    else:
        unidade_tempo = "minutos"
    
    # Exibe os resultados
    print(f"\nMelhor Lag para Variação da Temperatura da Água: {melhor_lag} períodos")
    print(f"Variação necessária na temperatura externa para alterar 1°C na água: {variacao_necessaria_agua:.2f}°C")
    print(f"Variação necessária na temperatura externa para alterar 1°C na interna: {variacao_necessaria_interna:.2f}°C")
    print(f"Tempo necessário para a alteração da temperatura externa refletir na água: {tempo_reflexao_agua:.2f} {unidade_tempo}")
    
    # Retorna os resultados
    return variacao_necessaria_agua, variacao_necessaria_interna, tempo_reflexao_agua, intervalo_medio, unidade_tempo

    # Exibe os resultados
    print(f"\nMelhor Lag para Variação da Temperatura da Água: {melhor_lag} períodos")
    print(f"Variação necessária na temperatura externa para alterar 1°C na água: {variacao_necessaria_agua:.2f}°C")
    print(f"Variação necessária na temperatura externa para alterar 1°C na interna: {variacao_necessaria_interna:.2f}°C")
    print(f"Tempo necessário para a alteração da temperatura externa refletir na água: {tempo_reflexao_agua:.2f} {unidade_tempo}")
    
    # Retorna os resultados
    return variacao_necessaria_agua, variacao_necessaria_interna, tempo_reflexao_agua, intervalo_medio, unidade_tempo
def realizar_analise(dados, sensores):
    limites = {
        'Temperatura Ar': {
            'minima': float(entry_limite_ar_min.get()),
            'maxima': float(entry_limite_ar_max.get())
        },
        'Temperatura Água': {
            'minima': float(entry_limite_agua_min.get()),
            'maxima': float(entry_limite_agua_max.get())
        },
        'Variação Água': {
            'maxima': float(entry_limite_variacao.get())
        }
    }

    # Avalia os dias com problemas nas condições de temperatura
    dias_problemas = avaliar_condicoes(dados, sensores, limites)
    
    # Chama a função avaliar_impacto_estufa com as devidas alterações
    variacao_necessaria_agua, variacao_necessaria_interna, tempo_reflexao_agua, intervalo_medio, unidade_tempo = avaliar_impacto_estufa(dados, sensores)
    # Criação da string para os resultados da análise
    resultados_texto = ""
    resultado_texto_previsao="\n Análise com base na previsão do tempo para os proximos 5 dias"
    resultados_texto += "\n===== DIAGNÓSTICO DA ESTUFA =====\n\n"
    resultados_texto += "Interpretação dos Resultados da Análise de Regressão:\n"
    
    # Exibe as variações necessárias
    resultados_texto += f"- Variação necessária na temperatura externa para alterar 1°C na água: {variacao_necessaria_agua:.2f}°C\n"
    resultados_texto += f"- Variação necessária na temperatura externa para alterar 1°C na temperatura interna: {variacao_necessaria_interna:.2f}°C\n"
    
    # Exibe o tempo de reflexão na água
    resultados_texto += f"- Tempo necessário para refletir a variação externa na temperatura da água: {tempo_reflexao_agua:.2f} {unidade_tempo}\n"
    cidade = cidade_var.get()
    #####################################################################
    #####################################################################
    
    
    
    
    # Atribuindo cada temperatura média a uma variável diferente
   
    
    # Chame a função
    status_api = testar_api_openweathermap(api_key="2fbe3c5a6c138d8ba0f037e7537e7a8d")

        
    if status_api == 1:
        print("A API está disponível para uso.")
        resultado_previsao = prever_comportamento_temperaturas(cidade, dados, sensores, API_KEY="2fbe3c5a6c138d8ba0f037e7537e7a8d")
        texto_previsao=avaliar_condicoes_previsao(resultado_previsao, limites)
        if texto_previsao:
            resultado_texto_previsao+="Dias com Problemas Detectados:\n\n"
            for dia, problemas in dias_problemas:
                resultado_texto_previsao += f"Data: {dia}\n"
                for problema in problemas:
                    resultado_texto_previsao += f"  - {problema}\n"
                resultado_texto_previsao += "\n"
    else:
        print("A API não está disponível. Verifique a conexão ou a chave da API.")
        plotar_grafico_previsao(pd.DataFrame())  # Passa um DataFrame vazio para exibir a mensagem

    resultados_texto += "\n===== DIAGNÓSTICO DA AMBIÊNCIA =====\n\n"
    
    # Verifica se houve problemas com as condições de temperatura
    if dias_problemas:
        resultados_texto += "Dias com Problemas Detectados:\n\n"
        for dia, problemas in dias_problemas:
            resultados_texto += f"Data: {dia}\n"
            for problema in problemas:
                resultados_texto += f"  - {problema}\n"
            resultados_texto += "\n"
    else:
        resultados_texto += "Nenhum problema encontrado nas condições de temperatura.\n"
    
    return resultados_texto, resultado_texto_previsao

def plotar_grafico(dados, sensores):
    """Cria um gráfico com os dados filtrados e adiciona um cursor interativo usando mplcursors."""
    datas_horas = [datetime.strptime(f"{dado['Data']} {dado['Hora']}", '%d/%m/%Y %H:%M:%S') for dado in dados]
    fig, ax1 = plt.subplots(figsize=(8, 4))  # Ajuste o tamanho do gráfico
    
    # Armazenar os handles das linhas para mplcursors
    handles = []
    labels = []
    
    for sensor in sensores:
        temperaturas = [float(dado[sensor]) for dado in dados]
        handle, = ax1.plot(datas_horas, temperaturas, label=sensor, marker='o')
        handles.append(handle)
        labels.append(sensor)
    
    ax1.set_xlabel('Data e Hora')
    ax1.set_ylabel('Temperatura (°C)')
    ax1.set_title('Temperaturas ao Longo do Período')
    ax1.legend()
    # Ajustar o intervalo do eixo X com base no tamanho dos dados
    num_pontos = len(datas_horas)
    if num_pontos > 24:
        intervalo = max(3, num_pontos // 24)  # Garante que o intervalo seja pelo menos 1
    else:
        intervalo = 3  # Se houver 26 ou menos pontos, mostra todos
    
    ax1.xaxis.set_major_locator(mdates.HourLocator(interval=intervalo))  # Reduz a quantidade de marcações para cada 12 horas
    ax1.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%y %H:%M'))  # Formato para dia e hora
    ax1.figure.autofmt_xdate()  # Rotaciona as datas para melhor visualização
    
    # Adiciona o cursor interativo
    mplcursors.cursor(handles, hover=True).connect(
        "add", lambda sel: sel.annotation.set_text(
            f"{sel.artist.get_label()}\n"
            f"Data: {mdates.num2date(sel.target[0]).strftime('%d/%m/%Y %H:%M')}\n"
            f"Temperatura (°C): {sel.target[1]:.2f}"
        )
    )
    
    # Ajusta o layout para evitar sobreposição
    fig.tight_layout()
    
    return fig
def exibir_temperaturas(sensor, dados, dia_max_amplitude):
    """Exibe as temperaturas máxima e mínima e suas horas."""
    temperaturas = [float(dado[sensor]) for dado in dados if dado['Data'] == dia_max_amplitude]
    horarios = [dado['Hora'] for dado in dados if dado['Data'] == dia_max_amplitude]
    
    temp_max = max(temperaturas)
    temp_min = min(temperaturas)
    hora_max = horarios[temperaturas.index(temp_max)]
    hora_min = horarios[temperaturas.index(temp_min)]
    
    return temp_max, hora_max, temp_min, hora_min
def atualizar_interface():
    """Atualiza a interface com base nas entradas do usuário."""
    formatos_data = ['%d-%m-%Y', '%d/%m/%Y']
    
    def validar_data(data_str):
        for formato in formatos_data:
            try:
                return datetime.strptime(data_str, formato)
            except ValueError:
                continue
        return None
    
    data_inicio = entry_data_inicio.get()
    data_fim = entry_data_fim.get()
    
    if data_inicio and data_fim:
        data_inicio_valida = validar_data(data_inicio)
        data_fim_valida = validar_data(data_fim)
        
        if data_inicio_valida and data_fim_valida:
            data_inicio = data_inicio_valida
            data_fim = data_fim_valida
        else:
            messagebox.showerror('Erro', 'Formato de data inválido. Use dd-mm-aaaa ou dd/mm/aaaa.')
            return
    else:
        data_inicio = None
        data_fim = None
    
    dados, cabecalho = ler_dados(caminho_arquivo)
    sensores = [sensor for sensor in cabecalho if sensor not in {'Data', 'Hora'}]
    
    if data_inicio and data_fim:
        dados_filtrados = filtrar_dados_por_periodo(dados, data_inicio.strftime('%d-%m-%Y'), data_fim.strftime('%d-%m-%Y'))
    else:
        dados_filtrados = dados
    
    if dados_filtrados:
        # Seleciona o sensor de temperatura da água (primeira posição)
        sensor_agua = sensores[0]
        
        # Calcula a amplitude térmica para o sensor de água e para todos os sensores
        dia_max_amplitude, dados_max_amplitude = calcular_amplitude_termica(dados_filtrados, sensores)
        picos_variacao = calcular_variacao(dados_filtrados, sensores, dia_max_amplitude)
        
        # Chama a função de realizar análise e atualiza os resultados no texto
        resultados, resultados_pr = realizar_analise(dados_filtrados, sensores)
        text_resultados_analise.config(state=tk.NORMAL)
        text_resultados_analise.delete(1.0, tk.END)  # Limpa o quadro de texto
        text_resultados_analise.insert(tk.END, resultados)  # Insere os novos resultados
        text_resultados_analise.config(state=tk.DISABLED)  # Desabilita o quadro de texto
        
        # Agora chamamos a função para obter estatísticas
        estatisticas = calcular_estatisticas_temperaturas(dados_filtrados, sensores)

        # Calcula o dia mais quente
        dia_mais_quente, resultados_quente = calcular_dia_mais_quente(dados_filtrados, sensores)
        
        # Calcula o dia mais frio
        dia_mais_frio, resultados_frio = calcular_dia_mais_frio(dados_filtrados, sensores)
       
        # Atualiza o quadro de resultados
        resultados = []
        
        resultados.append(f'Dia com a maior amplitude térmica para o sensor de água ({sensor_agua}): {dia_max_amplitude}\n')
        
        # Temperaturas máximas, mínimas e variação para cada sensor no dia com a maior amplitude do sensor de água
        for sensor, info in dados_max_amplitude.items():
            temp_max = info['temp_max']
            temp_min = info['temp_min']
            var = info['variacao']
            resultados.append(f'{sensor}: Máx = {temp_max} (°C), Mín = {temp_min} (°C), Variação = {var} (°C)\n')
        
        resultados.append('\nPico de variação entre as amostras:\n')
        for sensor, (variacao, tipo_variacao, hora) in picos_variacao.items():
            resultados.append(f'{sensor}: Variação = {variacao} (°C) ({tipo_variacao}) (Hora: {hora})\n')
        
        # Adiciona estatísticas das temperaturas
        resultados.append('\nEstatísticas das Temperaturas:\n')
        for sensor, estatisticas_sensor in estatisticas.items():
            resultados.append(f'\nEstatísticas para o sensor {sensor}:\n')
            resultados.append(f'- Média da Temperatura no Período: {estatisticas_sensor["media_temperatura_periodo"]:.2f} °C\n')
            resultados.append(f'- Média das Temperaturas Máximas Diárias: {estatisticas_sensor["media_temperatura_max"]:.2f} °C\n')
            resultados.append(f'- Média das Temperaturas Mínimas Diárias: {estatisticas_sensor["media_temperatura_min"]:.2f} °C\n')
            resultados.append(f'- Média da Amplitude Térmica: {estatisticas_sensor["media_amplitude"]:.2f} °C\n')
        
        # Adiciona informações do dia mais quente
        resultados.append(f'\nDia mais quente: {dia_mais_quente}\n')
        for sensor, info in resultados_quente.items():
            resultados.append(f'{sensor}: Máx = {info["temp_max"]} (°C), Mín = {info["temp_min"]} (°C), Variação = {info["variacao"]} (°C)\n')
        
        # Adiciona informações do dia mais frio
        resultados.append(f'\nDia mais frio: {dia_mais_frio}\n')
        for sensor, info in resultados_frio.items():
            resultados.append(f'{sensor}: Máx = {info["temp_max"]} (°C), Mín = {info["temp_min"]} (°C), Variação = {info["variacao"]} (°C)\n')
        
        
        text_resultados.config(state=tk.NORMAL)
        text_resultados.delete(1.0, tk.END)
        text_resultados.insert(tk.END, ''.join(resultados))
        text_resultados.config(state=tk.DISABLED)
        
        # Atualiza o gráfico
        fig = plotar_grafico(dados_filtrados, sensores)
        
        for widget in frame_grafico.winfo_children():
            widget.destroy()
        
        canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    else:
        messagebox.showinfo('Nenhum dado', 'Nenhum dado encontrado para o período especificado.')
    
def escolher_arquivo():
    """Abre um diálogo para escolher o arquivo de entrada.""" 
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
        filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
    )
    if caminho_arquivo:
        dados, cabecalho = ler_dados(caminho_arquivo)
        data_inicial, data_final = periodo_coleta(dados)
        label_periodo.config(text=f'Período disponível: {data_inicial} a {data_final}')
def salvar_configuracao():
    """Salva a configuração em um arquivo config.txt solicitado pelo usuário."""    
    try:
        # Coletar todos os valores
        intervalo = int(entry_intervalo.get())
        unidade = unidade_var.get()
        nome_arquivo = entry_nome_arquivo.get()
        wifi_ssid = entry_wifi_ssid.get()
        wifi_password = entry_wifi_password.get()
        
        # Converter intervalo para minutos
        intervalo_minutos = intervalo * 60 if unidade == 'Horas' else intervalo
        
        # Solicitar local para salvar
        caminho_arquivo = filedialog.asksaveasfilename(
            defaultextension='.txt',
            filetypes=[('Arquivos de Texto', '*.txt')],
            title='Salvar Configuração',
            initialfile='conf.txt'
        )
        
        if caminho_arquivo:
            with open(caminho_arquivo, 'w') as arquivo:
                arquivo.write(f'INTERVALO={intervalo_minutos}\n')
                arquivo.write(f'ARQUIVO={nome_arquivo}.txt\n')  # Garante extensão .txt
                arquivo.write(f'WIFI_SSID={wifi_ssid}\n')
                arquivo.write(f'WIFI_PASSWORD={wifi_password}\n')
            
            messagebox.showinfo('Configuração Salva', 'As configurações foram salvas com sucesso!')
        else:
            messagebox.showwarning('Aviso', 'Nenhum arquivo foi selecionado para salvar.')
    
    except ValueError:
        messagebox.showerror('Erro', 'Intervalo deve ser um número inteiro.')
def formatar_data(data_str):
    """Formata a data para o formato %d-%m-%Y."""
    try:
        # Tenta ler a data no formato atual
        data = datetime.strptime(data_str, '%d/%m/%Y')
    except ValueError:
        # Se a data não corresponder ao formato, tenta outros formatos
        data = datetime.strptime(data_str, '%d/%m/%Y')
    
    return data.strftime('%d-%m-%Y')
def adicionar_tabela(pdf, dados, cabecalho, titulo="Dados da Tabela"):
    """Adiciona uma tabela com os dados ao PDF."""
   # pdf.add_page()
    pdf.set_font('Arial', 'B', 12)
    
    # Adiciona o título da tabela
    pdf.cell(0, 10, titulo, ln=True, align='C')  # Título centralizado com espaçamento vertical
    
    pdf.set_font('Arial', 'B', 10)

    # Adiciona o cabeçalho da tabela
    col_width = pdf.get_string_width(max(cabecalho, key=len)) + 2
    for coluna in cabecalho:
        pdf.cell(col_width, 8, coluna, 1, 0, 'C')
    pdf.ln()

    # Adiciona as linhas da tabela
    pdf.set_font('Arial', '', 8)
    for linha in dados:
        for coluna in cabecalho:
            pdf.cell(col_width, 8, str(linha.get(coluna, '')), 1, 0, 'C')
        pdf.ln()
def converter_txt_para_xlsx(caminho_arquivo_txt, diretorio_saida):
    """Converte um arquivo .txt para .xlsx com formatação e salva no diretório especificado."""
    try:
        # Lê o arquivo .txt usando pandas
        df = pd.read_csv(caminho_arquivo_txt, delimiter='\t')  # Ajuste o delimitador conforme necessário
        
        # Exibir o cabeçalho do DataFrame para referência
        print("Cabeçalho do arquivo TXT:", df.columns.tolist())
        
        # Ajustar o DataFrame conforme esperado
        df.columns = ['Data', 'Hora', 'Temp_Sensor1', 'Temp_Sensor2', 'Temp_Sensor3']
        df['Data_Hora'] = pd.to_datetime(df['Data'] + ' ' + df['Hora'], dayfirst=True)
        df = df[['Data_Hora', 'Temp_Sensor1', 'Temp_Sensor2', 'Temp_Sensor3']]
        
        # Renomear as colunas para os nomes específicos dos sensores
        df.columns = ['Data_Hora', 'Temperatura da Água', 'Temperatura da Estufa', 'Temperatura Externa']
        
        # Define o caminho do arquivo XLSX
        nome_base = os.path.splitext(os.path.basename(caminho_arquivo_txt))[0]
        data_atual = datetime.now().strftime('%d-%m-%Y')
        caminho_arquivo_xlsx = os.path.join(diretorio_saida, f"{nome_base}_{data_atual}.xlsx")

        # Criar um novo Workbook e carregar a planilha
        wb = Workbook()
        ws = wb.active
        
        # Adicionar os dados ao Workbook
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        
        # Ajustar a largura das colunas
        column_widths = {
            'A': 25,  # Data_Hora
            'B': 25,  # Temperatura da Água
            'C': 25,  # Temperatura da Estufa
            'D': 25,  # Temperatura Externa
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Formatar o cabeçalho em negrito e com cor de preenchimento
        header_font = Font(bold=True, color='FFFFFF')  # Cor do texto em branco
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')  # Cor de fundo azul
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Salvar a planilha
        wb.save(caminho_arquivo_xlsx)
        
        #print(f"Arquivo salvo como: {caminho_arquivo_xlsx}")
        #messagebox.showinfo('Sucesso', f'Arquivo convertido e salvo como {caminho_arquivo_xlsx}!')
    
    except Exception as e:
        messagebox.showerror('Erro', f'Ocorreu um erro ao converter o arquivo TXT: {e}')
def salvar_relatorio():
    """Salva o gráfico e o relatório em um arquivo PDF e o XLSX no mesmo diretório."""
    try:
        if not caminho_arquivo:
            messagebox.showerror('Erro', 'Nenhum arquivo carregado.')
            return

        # Solicita o diretório para salvar o relatório PDF
        diretorio_relatorio = filedialog.askdirectory(title='Escolha o diretório para salvar o relatório')
        
        if not diretorio_relatorio:
            messagebox.showwarning('Aviso', 'Nenhum diretório foi selecionado.')
            return
        
        # Converte o arquivo TXT para XLSX e salva no mesmo diretório
        converter_txt_para_xlsx(caminho_arquivo, diretorio_relatorio)

        # Formatar as datas
        data_inicio_str = entry_data_inicio.get()
        data_fim_str = entry_data_fim.get()
        
        data_inicio_str_formatado = formatar_data(data_inicio_str)
        data_fim_str_formatado = formatar_data(data_fim_str)

        dados, cabecalho = ler_dados(caminho_arquivo)
        sensores = [sensor for sensor in cabecalho if sensor not in {'Data', 'Hora'}]
        
        if data_inicio_str_formatado and data_fim_str_formatado:
            dados_filtrados = filtrar_dados_por_periodo(dados, data_inicio_str_formatado, data_fim_str_formatado)
        else:
            dados_filtrados = dados
        
        if dados_filtrados:
            # Seleciona o sensor de temperatura da água (primeira posição)
            sensor_agua = sensores[0]

            # Calcula a amplitude térmica e variações
            dia_max_amplitude, dados_max_amplitude = calcular_amplitude_termica(dados_filtrados, sensores)
            picos_variacao = calcular_variacao(dados_filtrados, sensores, dia_max_amplitude)
            
            # Agora chamamos a função para obter estatísticas
            estatisticas = calcular_estatisticas_temperaturas(dados_filtrados, sensores)
            
            # Calcula o dia mais quente
            dia_mais_quente, resultados_quente = calcular_dia_mais_quente(dados_filtrados, sensores)
            
            # Calcula o dia mais frio
            dia_mais_frio, resultados_frio = calcular_dia_mais_frio(dados_filtrados, sensores)
            
            # Atualiza o relatório
            resultados = []
            resultadosI = []
            
            resultadosI.append(f'Dia com a maior amplitude térmica para o sensor de água ({sensor_agua}): {dia_max_amplitude}\n')
            
            # Temperaturas máximas, mínimas e variação para cada sensor no dia com a maior amplitude do sensor de água
            dados_max_amplitude_sensores = calcular_amplitude_termica(dados_filtrados, sensores)[1]
            for sensor, info in dados_max_amplitude_sensores.items():
                temp_max = info['temp_max']
                temp_min = info['temp_min']
                var = info['variacao']
                resultadosI.append(f'{sensor}: Máx = {temp_max} (°C), Mín = {temp_min} (°C), Variação = {var} (°C)\n')
            resultadosI.append('\nPico de variação:\n')
            for sensor, (variacao, tipo_variacao, hora) in picos_variacao.items():
                resultadosI.append(f'{sensor}: Variação = {variacao} (°C) ({tipo_variacao}) (Hora: {hora})\n')
            
            # Adiciona estatísticas das temperaturas
            resultados.append('\nInformações Gerais:\n')
            for sensor, estatisticas_sensor in estatisticas.items():
                resultados.append(f'\nEstatísticas para o sensor {sensor}:\n')
                resultados.append(f'- Média da Temperatura no Período: {estatisticas_sensor["media_temperatura_periodo"]:.2f} °C\n')
                resultados.append(f'- Média das Temperaturas Máximas Diárias: {estatisticas_sensor["media_temperatura_max"]:.2f} °C\n')
                resultados.append(f'- Média das Temperaturas Mínimas Diárias: {estatisticas_sensor["media_temperatura_min"]:.2f} °C\n')
                resultados.append(f'- Média da Amplitude Térmica: {estatisticas_sensor["media_amplitude"]:.2f} °C\n')

            
            # Adiciona informações do dia mais quente
            resultados.append(f'\nDia mais quente: {dia_mais_quente}\n')
            for sensor, info in resultados_quente.items():
                resultados.append(f'{sensor}: Máx = {info["temp_max"]} (°C), Mín = {info["temp_min"]} (°C), Variação = {info["variacao"]} (°C)\n')
            
            # Adiciona informações do dia mais frio
            resultados.append(f'\nDia mais frio: {dia_mais_frio}\n')
            for sensor, info in resultados_frio.items():
                resultados.append(f'{sensor}: Máx = {info["temp_max"]} (°C), Mín = {info["temp_min"]} (°C), Variação = {info["variacao"]} (°C)\n')
            
            
            
            # Salvar o gráfico geral do dia com maior amplitude térmica
            fig_max_amplitude = plt.figure(figsize=(15, 8 ))
            ax_max = fig_max_amplitude.add_subplot(111)
            datas_horas_max = [datetime.strptime(f"{dado['Data']} {dado['Hora']}", '%d/%m/%Y %H:%M:%S') for dado in dados if dado['Data'] == dia_max_amplitude]
            
            for sensor in sensores:
                temperaturas_max = [float(dado[sensor]) for dado in dados if dado['Data'] == dia_max_amplitude]
                ax_max.plot(datas_horas_max, temperaturas_max, label=sensor, marker='o')
            
            ax_max.set_title(f'Gráfico do Dia com Maior Amplitude Térmica ({dia_max_amplitude})')
            ax_max.set_xlabel('Data e Hora')
            ax_max.set_ylabel('Temperatura (°C)')
            ax_max.legend()
            ax_max.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
            fig_max_amplitude.autofmt_xdate()
            
            caminho_grafico_max = os.path.join(diretorio_relatorio, 'grafico_max_amplitude.png')
            fig_max_amplitude.savefig(caminho_grafico_max, format='png')
            plt.close(fig_max_amplitude)
            
            # Salvar os gráficos de cada sensor individualmente
            caminhos_graficos_sensores = []
            for sensor in sensores:
                fig_sensor = plt.figure(figsize=(15, 8))
                ax_sensor = fig_sensor.add_subplot(111)
                datas_horas = [datetime.strptime(f"{dado['Data']} {dado['Hora']}", '%d/%m/%Y %H:%M:%S') for dado in dados_filtrados]
                temperaturas = [float(dado[sensor]) for dado in dados_filtrados]
                ax_sensor.plot(datas_horas, temperaturas, label=sensor)
                ax_sensor.set_title(f'Gráfico de Temperatura - {sensor}')
                ax_sensor.set_xlabel('Data e Hora')
                ax_sensor.set_ylabel('Temperatura (°C)')
                ax_sensor.legend()
                # Ajustar o intervalo do eixo X com base no tamanho dos dados
                num_pontos = len(datas_horas)
                if num_pontos > 24:
                    intervalo = max(3, num_pontos // 24)  # Garante que o intervalo seja pelo menos 1
                else:
                    intervalo = 3  # Se houver 26 ou menos pontos, mostra todos
                
                
                ax_sensor.xaxis.set_major_locator(mdates.HourLocator(interval=intervalo))
                #ax_sensor.xaxis.set_major_locator(mdates.HourLocator(interval=12))
                ax_sensor.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%y'))
                fig_sensor.autofmt_xdate()
                
                caminho_grafico_sensor = os.path.join(diretorio_relatorio, f'grafico_{sensor}.png')
                fig_sensor.savefig(caminho_grafico_sensor, format='png')
                caminhos_graficos_sensores.append(caminho_grafico_sensor)
                
                plt.close(fig_sensor)
            # Criar texto para os limites
            cidade = cidade_var.get()
            limites = {
                'Temperatura Ar': {
                    'minima': float(entry_limite_ar_min.get()),
                    'maxima': float(entry_limite_ar_max.get())
                },
                'Temperatura Água': {
                    'minima': float(entry_limite_agua_min.get()),
                    'maxima': float(entry_limite_agua_max.get())
                },
                'Variação Água': {
                    'maxima': float(entry_limite_variacao.get())
                }
            }
            resultados_analise, resultado_texto_previ = realizar_analise(dados_filtrados, sensores)
            resultado_previsao = prever_comportamento_temperaturas(cidade, dados_filtrados, sensores, API_KEY="2fbe3c5a6c138d8ba0f037e7537e7a8d")
            
            texto_previsao=avaliar_condicoes_previsao(resultado_previsao, limites)
            resultado_texto_previsao="\n Análise com base na previsão do tempo para os proximos 5 dias\n "
            if texto_previsao:
                resultado_texto_previsao+="Dias com Problemas Detectados:\n\n"
                for dia, problemas in texto_previsao:
                    resultado_texto_previsao += f"Data: {dia}\n"
                    for problema in problemas:
                        resultado_texto_previsao += f"  - {problema}\n"
                    resultado_texto_previsao += "\n"
            # Integrar resultados da análise ao relatório
            #relatorio_texto = ''.join(resultados) + resultados_analise    
            
            
            text_resultados.config(state=tk.NORMAL)
            text_resultados.delete(1.0, tk.END)
            text_resultados.insert(tk.END, ''.join(resultados))
            text_resultados.config(state=tk.DISABLED)
            relatorio_textoII = ''.join(resultado_texto_previsao)
            relatorio_textoI = ''.join(resultadosI)
            relatorio_texto = ''.join(resultados)
            # Extrair o nome base do arquivo original e adicionar "relatorio" e a data atual
            nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
            data_atual = datetime.now().strftime('%d_%m_%Y')
            nome_relatorio = os.path.join(diretorio_relatorio, f"{nome_base}_relatorio_{data_atual}.pdf")
            
            
            
            
            
                        
            # Previsão para os próximos 5 dias - Chama a função para gerar o gráfico de previsão
            #resultado_previsao = prever_comportamento_temperaturas(cidade, dados, sensores, API_KEY="2fbe3c5a6c138d8ba0f037e7537e7a8d")
            
            # Salva o gráfico de previsão
            fig_previsao = plt.figure(figsize=(15, 8))
            ax_previsao = fig_previsao.add_subplot(111)
            ax_previsao.plot(resultado_previsao['DataHora'], resultado_previsao['TemperaturaExterna'], label='Temp. Externa', marker='o', color='blue')
            ax_previsao.plot(resultado_previsao['DataHora'], resultado_previsao['Prev_Agua'], label='Temp. Água', marker='o', color='green')
            ax_previsao.plot(resultado_previsao['DataHora'], resultado_previsao['Prev_Interna'], label='Temp. Interna', marker='o', color='red')
            
            ax_previsao.set_title('Previsão do Comportamento das Temperaturas nos Próximos 5 Dias')
            ax_previsao.set_xlabel('Data e Hora')
            ax_previsao.set_ylabel('Temperatura (°C)')
            ax_previsao.legend()
            ax_previsao.xaxis.set_major_formatter(mdates.DateFormatter('%d-%m-%y %H'))
            fig_previsao.autofmt_xdate()
            
            caminho_grafico_previsao = os.path.join(diretorio_relatorio, 'grafico_previsao_temperaturas.png')
            fig_previsao.savefig(caminho_grafico_previsao, format='png')
            plt.close(fig_previsao)
            
            
                        
                        
                        
                        
                        
                        
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            # Salvar o relatório em um arquivo PDF
            pdf = FPDF()
            pdf.add_page()
            altura_pagina = 297  # Altura de uma página A4 em mm
            # Adicionar o logo ao lado do título
            logo_path = 'englife.png'  # Caminho para a imagem do logotipo
            pdf.image(logo_path, x=10, y=8, w=30)  # Ajuste a posição e tamanho conforme necessário
            
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, 'Relatório de Análise de Dados de Temperatura', ln=True, align='C')
            pdf.ln(15)
            
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 8, relatorio_textoI)
            
            pdf.ln(5)
            
            
            
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, 'Gráfico Geral do Dia com Maior Amplitude:', ln=True, align='L')
            pdf.image(caminho_grafico_max, x=10, y=pdf.get_y() + 2, w=160)  # Ajustado para largura menor
            
        
            # Adicionar espaço extra entre o gráfico geral e os gráficos individuais
            pdf.ln(80)  # Espaço adicional de 30 unidades entre o gráfico geral e o bloco de gráficos individuais
            
            
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 8, relatorio_texto)
            
            pdf.ln(5)
            
            # Adicionar gráficos individuais dos sensores
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, 'Gráficos de Temperatura por Sensor:', ln=True, align='L')

            y_offset = pdf.get_y()
            for caminho_grafico_sensor in caminhos_graficos_sensores:
                if y_offset + 50 > 275:  # Ajusta o limite de acordo com a altura da página
                    pdf.add_page()
                    y_offset = 10
                pdf.image(caminho_grafico_sensor, x=10, y=y_offset, w=160)  # Ajustado para largura menor
                y_offset += 90  # Ajusta o espaço entre gráficos

            pdf.ln(170) 
            
            
                  
            # Adiciona a análise de ambiência logo em seguida, conforme solicitado
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, 'Análise Ambiência:', ln=True, align='L')
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 8, resultados_analise)
            
            '''
            # Mover para a próxima página se houver mais de 20 linhas ocupadas
            if  y_offset > 170:
                pdf.add_page()
            '''
            
            y_offset = pdf.get_y()  # Obtém a posição atual na página
            
            espaco_restante = altura_pagina - y_offset
            print(f"Espaço restante: {espaco_restante} mm")
            altura_conteudo = 100  # Exemplo de altura do conteúdo que você quer adicionar
            
            if espaco_restante < altura_conteudo:
                pdf.add_page()  # Adiciona uma nova página se não houver espaço suficiente

            # Adicionando o gráfico ao relatório PDF antes da análise de ambiência
            
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, 'Previsão para os próximos 5 dias:', ln=True, align='L')
            pdf.image(caminho_grafico_previsao, x=10, y=pdf.get_y() + 2, w=160)  # Ajustado para largura menor
            pdf.ln(90)  # Espaço extra após o gráfico
            
            
            # Limpar arquivos temporários (incluindo o novo gráfico)
            if os.path.exists(caminho_grafico_previsao):
                os.remove(caminho_grafico_previsao)
            pdf.set_font('Arial', '', 10)
            pdf.multi_cell(0, 8,  relatorio_textoII)
            # Adicionar a tabela logo após todos os gráficos
            pdf.ln(10)  # Espaço entre os gráficos e a tabela
            adicionar_tabela(pdf, dados, cabecalho)
            
            pdf.output(nome_relatorio)
            
            # Limpar arquivos temporários
            for caminho_grafico_sensor in caminhos_graficos_sensores:
                if os.path.exists(caminho_grafico_sensor):
                    os.remove(caminho_grafico_sensor)
            if os.path.exists(caminho_grafico_max):
                os.remove(caminho_grafico_max)
            
            messagebox.showinfo('Sucesso', f'Relatório salvo em: {nome_relatorio}')
        else:
            messagebox.showerror('Erro', 'Não há dados para gerar o relatório.')
    except Exception as e:
        messagebox.showerror('Erro', f'Ocorreu um erro ao salvar o relatório: {e}')
        
        
        
def atualizar_calendarios(periodo_inicio_str, periodo_fim_str):
    """Atualiza os calendários para refletir o período disponível e define as datas padrão."""
    # Converte as strings de data para objetos datetime.date
    periodo_inicio = datetime.strptime(periodo_inicio_str, '%d/%m/%Y').date()
    periodo_fim = datetime.strptime(periodo_fim_str, '%d/%m/%Y').date()
    
    # Atualiza os calendários para refletir o período disponível
    entry_data_inicio.config(mindate=periodo_inicio, maxdate=periodo_fim)
    entry_data_fim.config(mindate=periodo_inicio, maxdate=periodo_fim)
    
    # Define as datas padrão como o início e o fim do período disponível
    entry_data_inicio.set_date(periodo_inicio)
    entry_data_fim.set_date(periodo_fim)
def escolher_arquivo_e_atualizar():
    """Escolhe o arquivo e atualiza a interface e os calendários."""
    escolher_arquivo()
    dados, cabecalho = ler_dados(caminho_arquivo)
    periodo_inicio, periodo_fim = periodo_coleta(dados)
    atualizar_calendarios(periodo_inicio, periodo_fim)
    atualizar_interface()
def main():
    root = tk.Tk()
    root.title('Exemplo DateEntry')

    # Cria um DateEntry com o formato de data desejado
    date_entry = DateEntry(root, date_pattern='dd/mm/yyyy')
    date_entry.pack()

    # Exemplo de uso do Babel
    today = datetime.date.today()
    formatted_date = format_date(today, format='dd/MM/yyyy')
    print(f"Data formatada: {formatted_date}")

    root.mainloop()


def prever_comportamento_temperaturas(cidade, dados, sensores, API_KEY="2fbe3c5a6c138d8ba0f037e7537e7a8d"):
    # Chama a função para calcular as estatísticas
    estatisticas = calcular_estatisticas_temperaturas(dados, sensores)

    # Chama a função para avaliar o impacto da estufa
    variacao_necessaria_agua, variacao_necessaria_interna, tempo_reflexao_agua, intervalo_medio, unidade_tempo = avaliar_impacto_estufa(dados, sensores)

    # Atribuindo as temperaturas médias das estatísticas a variáveis iniciais
    temperatura_agua_atual = estatisticas['Temperatura da Água']['media_temperatura_periodo']
    temperatura_interna_atual = estatisticas['Temperatura da Estufa']['media_temperatura_periodo']

    # Testa se a API está disponível
    status_api = testar_api_openweathermap(API_KEY)

    if status_api == 1:
        print("A API está disponível para uso.")
        
        # Obtém as previsões climáticas da API
        link_previsao = f"https://api.openweathermap.org/data/2.5/forecast?q={cidade}&appid={API_KEY}&lang=pt_br"
        resposta = requests.get(link_previsao)
        dados_api = resposta.json()

        previsoes = []
        for item in dados_api['list']:
            temperatura_externa = item['main']['temp'] - 273.15  # Converter de Kelvin para Celsius
            data_hora = item['dt_txt']

            # Cálculo da variação da água de forma incremental
            if temperatura_externa > temperatura_agua_atual:
                variacao_agua = (temperatura_externa - temperatura_agua_atual) / variacao_necessaria_agua
            else:
                variacao_agua = (temperatura_externa - temperatura_agua_atual) / variacao_necessaria_agua

            # Atualiza a temperatura da água com base na variação
            temperatura_agua_atual += variacao_agua

            # Cálculo da variação da estufa de forma incremental
            if temperatura_externa > temperatura_interna_atual:
                variacao_interna = (temperatura_externa - temperatura_interna_atual) / variacao_necessaria_interna
            else:
                variacao_interna = (temperatura_externa - temperatura_interna_atual) / variacao_necessaria_interna

            # Atualiza a temperatura da estufa com base na variação
            temperatura_interna_atual += variacao_interna

            previsoes.append({
                'DataHora': data_hora,
                'TemperaturaExterna': temperatura_externa,
                'Prev_Agua': temperatura_agua_atual,
                'Prev_Interna': temperatura_interna_atual
            })

        # Converte a lista de previsões em DataFrame
        df_previsao = pd.DataFrame(previsoes)

        # Remove linhas com valores NaN resultantes do deslocamento
        df_previsao = df_previsao.dropna()

        # Plota o gráfico das previsões
        plotar_grafico_previsao(df_previsao)

        return df_previsao

    else:
        print("A API não está disponível. Verifique a conexão ou a chave da API.")
        # Plota um gráfico vazio
        plotar_grafico_previsao(pd.DataFrame())
        return pd.DataFrame()








def testar_api_openweathermap(api_key):
    link = f"http://api.openweathermap.org/data/2.5/weather?q=São Paulo&appid={api_key}"
    try:
        resposta = requests.get(link)
        resposta.raise_for_status()  # Levanta um erro se a resposta não for 200
        return 1  # API funcionando
    except requests.exceptions.RequestException:
        return 0  # API não acessível

# Exemplo de uso
cid = "São Paulo"
status_api = testar_api_openweathermap(cid)

if status_api == 1:
    print("A API está disponível para uso.")
else:
    print("A API não está disponível. Verifique a conexão ou a chave da API.")

def plotar_grafico_previsao(df_previsao):
    # Verifica se o DataFrame está vazio
    if df_previsao.empty:
        # Exibir mensagem na interface
        ax = figura.add_subplot(111)  # Adiciona um subplot na figura
        ax.text(0.5, 0.5, 'Nenhum dado disponível para plotar. A API pode não estar acessível.', 
                fontsize=12, ha='center', va='center', transform=ax.transAxes)
        ax.axis('off')  # Remove os eixos
        canvas.draw()  # Atualiza o canvas
        return  # Sai da função se não houver dados

    figura.clf()  # Limpa a figura atual
    ax = figura.add_subplot(111)  # Adiciona um subplot na figura
    
    # Plota as previsões e armazena os handles (linhas) para o cursor
    linha_externa, = ax.plot(df_previsao['DataHora'], df_previsao['TemperaturaExterna'], 
                             label='Temperatura Externa Prevista', color='red', linestyle='-', marker='o')
    linha_agua, = ax.plot(df_previsao['DataHora'], df_previsao['Prev_Agua'], 
                          label='Previsão da Temperatura da Água', color='blue', linestyle='--', marker='o')
    linha_interna, = ax.plot(df_previsao['DataHora'], df_previsao['Prev_Interna'], 
                             label='Previsão da Temperatura da Estufa', color='orange', linestyle='--', marker='o')

    ax.set_title('Previsão de Temperatura da Água, da Estufa e Temperatura Externa')
    ax.set_xlabel('Data e Hora')
    ax.set_ylabel('Temperatura (°C)')
    
    # Definindo rótulos do eixo x
    num_pontos = len(df_previsao)
    if num_pontos > 10:
        intervalo = num_pontos // 10  # Exibir 10 rótulos
        ax.set_xticks(range(0, num_pontos, max(1, intervalo)))
    else:
        ax.set_xticks(range(num_pontos))  # Exibir todos os rótulos se for menor ou igual a 10

    ax.tick_params(axis='x', rotation=25)  # Rotaciona os rótulos do eixo x
    
    # Ajustando o tamanho da fonte da legenda
    ax.legend(fontsize='small')  # Ou use um número, como fontsize=8

    plt.tight_layout()  # Ajusta o layout da plotagem
    
    # Adiciona o cursor interativo (com os handles das linhas)
    handles = [linha_externa, linha_agua, linha_interna]
    mplcursors.cursor(handles, hover=True).connect(
        "add", lambda sel: sel.annotation.set_text(
            f"{sel.artist.get_label()}\n"
            #f"Data: {sel.target[0]:.2f}\n"
            f"Temperatura (°C): {sel.target[1]:.2f}"
        )
    )

    # Desenha o gráfico na interface
    canvas.draw()

def avaliar_condicoes_previsao(dados_previsao, limites):
    """
    Avalia as condições de temperatura para cada dia da previsão e retorna os dias em que as condições não foram atendidas
    junto com os parâmetros específicos que estão fora dos limites, incluindo uma breve explicação sobre o problema.

    Parâmetros:
    - dados_previsao: DataFrame contendo 'DataHora', 'Prev_Agua' e 'Prev_Interna'.
    - limites: dicionário com os limites para avaliação das temperaturas.

    Retorna:
    - Lista de tuplas contendo a data e os motivos (parâmetros fora dos limites).
    """
    
    # Organizando as temperaturas por data
    temperaturas_por_data = defaultdict(lambda: {'Prev_Agua': [], 'Prev_Interna': []})
    
    for index, row in dados_previsao.iterrows():
        data = datetime.strptime(row['DataHora'].split()[0], '%Y-%m-%d').strftime('%d/%m/%Y')
        temperaturas_por_data[data]['Prev_Agua'].append(row['Prev_Agua'])
        temperaturas_por_data[data]['Prev_Interna'].append(row['Prev_Interna'])

    # Armazenando os dias em que as condições não foram atendidas
    condicoes_nao_atendidas = []

    # Avaliando as condições para cada dia
    for data, temperaturas in temperaturas_por_data.items():
        problema = []
        
        # Análise da temperatura da água
        temp_agua = temperaturas['Prev_Agua']
        media_temp_agua = sum(temp_agua) / len(temp_agua)
        variacao_agua = max(temp_agua) - min(temp_agua)

        if media_temp_agua < limites['Temperatura Água']['minima']:
            problema.append(f"Temperatura da Água abaixo do mínimo permitido (Média do dia: {media_temp_agua:.2f}°C, Mínima recomendada: {limites['Temperatura Água']['minima']}°C).\n"
                            "A temperatura da água está muito baixa, o que pode impactar a atividade e o metabolismo dos peixes.")
        if media_temp_agua > limites['Temperatura Água']['maxima']:
            problema.append(f"Temperatura da Água acima do máximo permitido (Média do dia: {media_temp_agua:.2f}°C, Máxima recomendada: {limites['Temperatura Água']['maxima']}°C).\n"
                            "A temperatura elevada da água pode causar estresse térmico nos peixes.")
        if variacao_agua > limites['Variação Água']['maxima']:
            problema.append(f"Variação da Temperatura da Água acima do limite permitido (Variação do dia: {variacao_agua:.2f}°C, Variação Máxima recomendada: {limites['Variação Água']['maxima']}°C).\n"
                            "A temperatura está instável, o que pode estressar os animais.")

        # Análise da temperatura da estufa
        temp_estufa = temperaturas['Prev_Interna']
        media_temp_estufa = sum(temp_estufa) / len(temp_estufa)

        if media_temp_estufa < limites['Temperatura Ar']['minima']:
            problema.append(f"Temperatura da Estufa abaixo do mínimo permitido (Média do dia: {media_temp_estufa:.2f}°C, Mínima recomendada: {limites['Temperatura Ar']['minima']}°C).\n"
                            "A temperatura da estufa está muito baixa, o que pode causar desconforto térmico e estresse.")
        elif media_temp_estufa > limites['Temperatura Ar']['maxima']:
            problema.append(f"Temperatura da Estufa acima do máximo permitido (Média do dia: {media_temp_estufa:.2f}°C, Máxima recomendada: {limites['Temperatura Ar']['maxima']}°C).\n"
                            "A temperatura da estufa está muito alta, podendo resultar em superaquecimento.")

        # Se houver problemas, armazenamos a data e os motivos
        if problema:
            condicoes_nao_atendidas.append((data, problema))

    # Exibe os resultados na label
    texto_resultados.config(state=tk.NORMAL)
    texto_resultados.delete(1.0, tk.END)  # Limpa o conteúdo anterior
    if condicoes_nao_atendidas:
        resultados = [f"Data: {data}, Problemas:\n" + "".join(problemas) for data, problemas in condicoes_nao_atendidas]
        texto_resultados.insert(tk.END, "\n".join(resultados))  # Insere novos resultados
    else:
        texto_resultados.insert(tk.END, "Todas as condições atendidas.")
    texto_resultados.config(state=tk.DISABLED)

        
    return condicoes_nao_atendidas




# Cria a janela principal
janela = tk.Tk()
janela.title('Análise de Dados de Temperatura')

# Define o ícone da janela (ícone .ico)
janela.iconbitmap('englife.ico')

# Formatando uma data para o formato longo em português do Brasil
data_formatada = format_date(datetime.today(), format='long', locale='pt_BR')
print("Data formatada:", data_formatada)

# Inicializa o caminho do arquivo
caminho_arquivo = ''

# Cria o Notebook (abas)
notebook = ttk.Notebook(janela)
notebook.pack(pady=10, fill=tk.BOTH, expand=True)

# Aba de Dados
aba_dados = ttk.Frame(notebook)
notebook.add(aba_dados, text='Dados')

# Frame para os controles de entrada (datas e botões)
frame_controles = tk.Frame(aba_dados)
frame_controles.grid(row=0, column=0, sticky='nsew')
wi=40
# Botão para escolher o arquivo
botao_abrir_arquivo = tk.Button(frame_controles, text='Abrir Arquivo', command=escolher_arquivo_e_atualizar, width=wi)
botao_abrir_arquivo.grid(row=0, column=0, columnspan=2, pady=5, sticky='w')
#botao_abrir_arquivo.grid(row=0, column=0, columnspan=2, pady=5)

# Label com o período disponível
label_periodo = tk.Label(frame_controles, text='Período disponível: ')
label_periodo.grid(row=1, column=0, columnspan=2, sticky='w', padx=5)

# Entradas de data usando DateEntry do tkcalendar
label_data_inicio = tk.Label(frame_controles, text='Início do período:')
label_data_inicio.grid(row=2, column=0, padx=5, sticky='w')
entry_data_inicio = DateEntry(frame_controles, width=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
entry_data_inicio.grid(row=2, column=1, padx=5, sticky='ew')

label_data_fim = tk.Label(frame_controles, text='Final do período:')
label_data_fim.grid(row=3, column=0, padx=5, sticky='w')
entry_data_fim = DateEntry(frame_controles, width=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='dd/mm/yyyy')
entry_data_fim.grid(row=3, column=1, padx=5, sticky='ew')

# Botão para atualizar a interface
botao_atualizar = tk.Button(frame_controles, text='Atualizar', command=atualizar_interface, width=wi)
botao_atualizar.grid(row=4, column=0, columnspan=2, pady=5, sticky='w')
#botao_atualizar.grid(row=4, column=0, columnspan=2, pady=5)
# Botão para salvar o relatório
botao_salvar_relatorio = tk.Button(frame_controles, text='Salvar Relatório', command=salvar_relatorio, width=wi)
botao_salvar_relatorio.grid(row=5, column=0, columnspan=2, pady=5, sticky='w')
#botao_salvar_relatorio.grid(row=5, column=0, columnspan=2, pady=5)

# Frame para o gráfico
frame_grafico = tk.Frame(aba_dados, bg='white', relief=tk.SUNKEN, borderwidth=1)
frame_grafico.grid(row=1, column=0, columnspan=3, sticky='nsew', padx=5, pady=5)

# Canvas para o gráfico
canvas_grafico = tk.Canvas(frame_grafico, bg='lightgray', width=200, height=150)  # Ajuste o tamanho conforme necessário
canvas_grafico.pack(fill=tk.BOTH, expand=True)

# Frame para o quadro de texto e título com barra de rolagem
frame_resultados = tk.Frame(aba_dados)
frame_resultados.grid(row=0, column=1, columnspan=2, sticky='nsew', padx=5, pady=5)

# Título do quadro de análise
titulo_resultados = tk.Label(frame_resultados, text='Análise de Amplitude e Variação', font=('Arial', 12, 'bold'))
titulo_resultados.pack(anchor='w')

# Adicionando barra de rolagem
scrollbar_resultados = tk.Scrollbar(frame_resultados)
scrollbar_resultados.pack(side=tk.RIGHT, fill=tk.Y)

# Quadro de texto para os resultados
text_resultados = tk.Text(frame_resultados, wrap='word', height=20, width=150, state='disabled', yscrollcommand=scrollbar_resultados.set)
text_resultados.pack(fill=tk.BOTH, expand=True)
scrollbar_resultados.config(command=text_resultados.yview)

# Configurações de estilo para títulos em negrito
text_resultados.tag_configure('titulo', font=('Arial', 10, 'bold'))

# Ajusta as proporções das colunas e linhas
aba_dados.grid_columnconfigure(0, weight=0)   # Coluna 0 não se expandirá
aba_dados.grid_columnconfigure(1, weight=1)   # Coluna 1 expandirá
aba_dados.grid_columnconfigure(2, weight=1)   # Coluna 2 expandirá
aba_dados.grid_rowconfigure(0, weight=1)      # Linha 0 expandirá
aba_dados.grid_rowconfigure(1, weight=1)      # Linha 1 expandirá

# Cria a aba para análise
aba_analise = ttk.Frame(notebook)
notebook.add(aba_analise, text='Análise da Ambiência')

## Frame para os resultados da análise
frame_analise = tk.Frame(aba_analise)
frame_analise.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

# Título do quadro de resultados
titulo_analise = tk.Label(frame_analise, text='Resultados da Análise', font=('Arial', 12, 'bold'))
titulo_analise.grid(row=0, column=1, sticky='w')

# Adicionando barra de rolagem para resultados
scrollbar_analise = tk.Scrollbar(frame_analise)
scrollbar_analise.grid(row=1, column=2, sticky='ns')

# Quadro de texto para os resultados
text_resultados_analise = tk.Text(frame_analise, wrap='word', height=20, width=60, state='disabled', yscrollcommand=scrollbar_analise.set)
text_resultados_analise.grid(row=1, column=1, sticky='nsew')
scrollbar_analise.config(command=text_resultados_analise.yview)

titulo_analise = tk.Label(frame_analise, text='Parâmetros de Ambiência', font=('Arial', 12, 'bold'))
titulo_analise.grid(row=0, column=0, sticky='w')
# Quadro para parâmetros das análises (limites)
frame_parametros = tk.Frame(frame_analise, borderwidth=2, relief=tk.GROOVE)
frame_parametros.grid(row=1, column=0, padx=5, pady=5, sticky='nsew')

# Limites de Temperatura do Ar
label_limite_ar_min = tk.Label(frame_parametros, text='Limite Mínimo Temperatura Ar (°C):')
label_limite_ar_min.grid(row=0, column=0, padx=5, sticky='w')
entry_limite_ar_min = tk.Entry(frame_parametros, width=20)
entry_limite_ar_min.insert(0, '18.41')  # Valor sugerido
entry_limite_ar_min.grid(row=0, column=1, padx=5, sticky='w')

label_limite_ar_max = tk.Label(frame_parametros, text='Limite Máximo Temperatura Ar (°C):')
label_limite_ar_max.grid(row=1, column=0, padx=5, sticky='w')
entry_limite_ar_max = tk.Entry(frame_parametros, width=20)
entry_limite_ar_max.insert(0, '30.7')  # Valor sugerido
entry_limite_ar_max.grid(row=1, column=1, padx=5, sticky='w')

# Limites de Temperatura da Água
label_limite_agua_min = tk.Label(frame_parametros, text='Limite Mínimo Temperatura Água (°C):')
label_limite_agua_min.grid(row=2, column=0, padx=5, sticky='w')
entry_limite_agua_min = tk.Entry(frame_parametros, width=20)
entry_limite_agua_min.insert(0, '28.17')  # Valor sugerido
entry_limite_agua_min.grid(row=2, column=1, padx=5, sticky='w')

label_limite_agua_max = tk.Label(frame_parametros, text='Limite Máximo Temperatura Água (°C):')
label_limite_agua_max.grid(row=3, column=0, padx=5, sticky='w')
entry_limite_agua_max = tk.Entry(frame_parametros, width=20)
entry_limite_agua_max.insert(0, '30.39')  # Valor sugerido
entry_limite_agua_max.grid(row=3, column=1, padx=5, sticky='w')

# Limites de Variação da Água
label_limite_variacao = tk.Label(frame_parametros, text='Limite aceitável da Variação na  Água (°C):')
label_limite_variacao.grid(row=4, column=0, padx=5, sticky='w')
entry_limite_variacao = tk.Entry(frame_parametros, width=20)
entry_limite_variacao.insert(0, '2.51')  # Valor sugerido
entry_limite_variacao.grid(row=4, column=1, padx=5, sticky='w')

# Botão para atualizar a interface
botao_atualizar = tk.Button(frame_parametros, text='Atualizar Interface', command=atualizar_interface,width=wi)
botao_atualizar.grid(row=5, columnspan=2, pady=5)

# Botão para salvar o relatório
botao_salvar_relatorio = tk.Button(frame_parametros, text='Salvar Relatório', command=salvar_relatorio, width=wi)
botao_salvar_relatorio.grid(row=6,columnspan=2, pady=10)


# Ajuste do layout para expandir corretamente
frame_parametros.grid_columnconfigure(0, weight=1)
frame_parametros.grid_columnconfigure(1, weight=1)

# Ajuste do frame_analise para expandir o quadro de resultados
frame_analise.grid_columnconfigure(1, weight=1)  # Coluna dos resultados se expande
frame_analise.grid_rowconfigure(1, weight=1)     # Linha dos resultados se expande


# Cria a aba para análise
aba_previsao = ttk.Frame(notebook)
notebook.add(aba_previsao, text='Análise Preditiva')

# Lista de cidades de São Paulo
cidades_sp = [
    'Araçatuba',
    'Araras',
    'Campinas',
    'Jaguariúna',
    'Pirassununga',
    'Rubinéia',
    'Santa Fé do Sul',
    'Santos',
    'São Bernardo do Campo',
    'São Paulo',
    'Sorocaba',
    'Osasco',
    'Itu',
    'Indaiatuba',
    'Mogi das Cruzes',
    'Botucatu',
    'Piratininga',
    'Pindamonhangaba',
    'Lins'
]

# Variável para armazenar a cidade escolhida
cidade_var = tk.StringVar(value='Jaguariúna')  # Pre-seleciona Pirassununga

# Frame para organizar a caixa de seleção e os botões na horizontal
frame_selecao_botoes = ttk.Frame(aba_previsao)
frame_selecao_botoes.pack(pady=10)

# Caixa de seleção para cidades (à esquerda)
cidade_selecao = ttk.Combobox(frame_selecao_botoes, textvariable=cidade_var, values=cidades_sp, state='readonly')
cidade_selecao.pack(side=tk.LEFT, padx=10)

# Botão para prever comportamento das temperaturas (ao centro)
botao_prever = ttk.Button(frame_selecao_botoes, text='Previsão para os próximos 5 dias', command=atualizar_interface)
botao_prever.pack(side=tk.LEFT, padx=10)

# Novo botão à direita
novo_botao = ttk.Button(frame_selecao_botoes, text='Salvar Relatório', command=salvar_relatorio, width=wi)  # Substitua 'outra_funcao' pela função desejada
novo_botao.pack(side=tk.LEFT, padx=10)



# Área para mostrar os resultados da previsão
texto_resultados = tk.Text(aba_previsao, height=10, width=60, state=tk.DISABLED)
texto_resultados.pack(fill=tk.BOTH, expand=True, pady=10)  # Ajuste aqui

# Área para o gráfico
figura = plt.figure(figsize=(8, 4))
canvas = FigureCanvasTkAgg(figura, master=aba_previsao)
canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)  # Ajuste aqui




# Aba de Configuração
aba_configuracao = ttk.Frame(notebook)
notebook.add(aba_configuracao, text='Configuração de Aquisição')



# Quadro para configuração do arquivo e intervalo
frame_configuracao = tk.Frame(aba_configuracao, borderwidth=2, relief=tk.GROOVE)
frame_configuracao.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

# Outros campos de configuração
label_intervalo = tk.Label(frame_configuracao, text='Intervalo de Coleta:')
label_intervalo.grid(row=0, column=0, padx=5, sticky='w')
entry_intervalo = tk.Entry(frame_configuracao, width=20)
entry_intervalo.grid(row=0, column=1, padx=5, sticky='w')

label_unidade = tk.Label(frame_configuracao, text='Unidade:')
label_unidade.grid(row=1, column=0, padx=5, sticky='w')
unidade_var = tk.StringVar(value='Horas')
radio_horas = tk.Radiobutton(frame_configuracao, text='Horas', variable=unidade_var, value='Horas')
radio_horas.grid(row=1, column=1, padx=5, sticky='w')
radio_minutos = tk.Radiobutton(frame_configuracao, text='Minutos', variable=unidade_var, value='Minutos')
radio_minutos.grid(row=1, column=2, padx=5, sticky='w')

label_nome_arquivo = tk.Label(frame_configuracao, text='Nome do Arquivo:')
label_nome_arquivo.grid(row=2, column=0, padx=5, sticky='w')
entry_nome_arquivo = tk.Entry(frame_configuracao, width=20)
entry_nome_arquivo.grid(row=2, column=1, padx=5, sticky='w')
# Dentro do frame_configuracao (Aba de Configuração de Aquisição), adicione:

# Configurações de Wi-Fi
label_wifi_ssid = tk.Label(frame_configuracao, text='Nome da rede Wi-Fi:')
label_wifi_ssid.grid(row=4, column=0, padx=5, sticky='w')
entry_wifi_ssid = tk.Entry(frame_configuracao, width=20)
entry_wifi_ssid.grid(row=4, column=1, padx=5, sticky='w')

label_wifi_password = tk.Label(frame_configuracao, text='Senha da rede:')
label_wifi_password.grid(row=5, column=0, padx=5, sticky='w')
entry_wifi_password = tk.Entry(frame_configuracao, width=20, show="*")  # Mostrar asteriscos para senha
entry_wifi_password.grid(row=5, column=1, padx=5, sticky='w')

# Ajuste a posição do botão de salvar para a linha 6

# Botão para salvar as configurações
botao_salvar = tk.Button(frame_configuracao, text='Salvar Configuração', command=salvar_configuracao, width=20)
botao_salvar.grid(row=6, column=0, columnspan=2, pady=10)
#botao_salvar.grid(row=3, column=0, columnspan=2, pady=10)


# Cria a aba de Ajuda
aba_ajuda = ttk.Frame(notebook)
notebook.add(aba_ajuda, text='Ajuda')

# Frame para a ajuda
frame_ajuda = tk.Frame(aba_ajuda)
frame_ajuda.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

# Título do quadro de ajuda
titulo_ajuda = tk.Label(frame_ajuda, text='Como Usar o Aplicativo', font=('Arial', 12, 'bold'))
titulo_ajuda.pack(anchor='w')

# Quadro de texto para a ajuda com barra de rolagem
scrollbar_ajuda = tk.Scrollbar(frame_ajuda)
scrollbar_ajuda.pack(side=tk.RIGHT, fill=tk.Y)

text_ajuda = tk.Text(frame_ajuda, wrap='word', height=20, width=60, state='normal', yscrollcommand=scrollbar_ajuda.set)
text_ajuda.pack(fill=tk.BOTH, expand=True)
scrollbar_ajuda.config(command=text_ajuda.yview)

# Passo a passo da ajuda
ajuda_texto = (
    "1. Clique em 'Abrir Arquivo' para selecionar o arquivo de dados.\n"
    "2. Os dados disponíveis no arquivo serão exibidos na aba 'Dados'.\n"
    "3. Utilize as entradas 'Início do período' e 'Final do período' para definir um intervalo específico.\n"
    "4. Clique em 'Atualizar' para visualizar os dados do intervalo selecionado.\n"
    "5. A análise de temperatura e variação térmica será apresentada na aba 'Análise da Ambiência'.\n"
    "6. Para salvar a análise, clique em 'Salvar Relatório'.\n"
    "7. Para ajustar as configurações de limites de temperatura, vá para a aba 'Análise da Ambiência'.\n"
    "8. Insira os limites desejados e clique em 'Atualizar'.\n"
    "9. Para criar um arquivo de configuração de aquisição preencha os campos e clique em 'Salvar configurações'.\n"
    "10. Para mais informações, consulte a documentação do aplicativo ou entre em contato com o suporte."
)

text_ajuda.insert(tk.END, ajuda_texto)
text_ajuda.config(state='disabled')  # Desabilitar edição do texto de ajuda

# Ajusta as proporções das colunas e linhas da aba de ajuda
aba_ajuda.grid_columnconfigure(0, weight=1)   # Coluna 0 expandirá
aba_ajuda.grid_rowconfigure(0, weight=1)      # Linha 0 expandirá



# Inicia a aplicação
janela.mainloop()
